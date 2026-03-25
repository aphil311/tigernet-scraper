"""Exporter module for TigerNet scraper."""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import AlumRecord


def write_excel(records: list[AlumRecord], path: str) -> str:
    """Export records to Excel file, return final path."""
    # Ensure directory exists
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Alumni"

    # Define headers
    headers = [
        "Name",
        "Class Year",
        "Degree",
        "Company",
        "Title",
        "Org",
        "Email",
        "LinkedIn",
        "Location",
    ]

    # Write header row
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data rows
    for row_idx, record in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=record.name)
        ws.cell(row=row_idx, column=2, value=record.class_year)
        ws.cell(row=row_idx, column=3, value=record.degree)
        ws.cell(row=row_idx, column=4, value=record.company)
        ws.cell(row=row_idx, column=5, value=record.title)
        ws.cell(row=row_idx, column=6, value=record.org)
        ws.cell(row=row_idx, column=7, value=record.email)
        ws.cell(row=row_idx, column=8, value=record.linkedin)
        ws.cell(row=row_idx, column=9, value=record.location)

    # Auto-size columns (approximate based on max length)
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(path)
    return os.path.abspath(path)


def write_timestamped_excel(
    records: list[AlumRecord], company: str, output_dir: str = "output"
) -> str:
    """Write records to Excel with timestamped filename: YYMMDD_HHMM_Company.xlsx."""
    timestamp = datetime.now().strftime("%y%m%d_%H%M")
    # Sanitize company name for filename (remove illegal chars, replace spaces)
    safe_company = "".join(
        c for c in company if c.isalnum() or c in (" ", "-", "_")
    ).rstrip()
    safe_company = safe_company.replace(" ", "_")
    filename = f"{timestamp}_{safe_company}.xlsx"
    path = os.path.join(output_dir, filename)
    return write_excel(records, path)
